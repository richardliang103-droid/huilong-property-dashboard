import json
import os
import sys
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook

# Load configuration from environment with sensible defaults
ROOT = Path(__file__).resolve().parents[1]
SOURCE = Path(os.getenv(
    "HUILONG_EXCEL",
    Path.home() / 'Library/CloudStorage/SynologyDrive-Hermes/Houses/迴龍物件追蹤.xlsx'
))
STATUS_SOURCE = Path(os.getenv(
    "HUILONG_STATUS_JSON",
    Path.home() / '.hermes/profiles/argus/data/huilong_watch_status.json'
))
DEST = ROOT / 'data/properties.json'

def clean(value):
    return value.isoformat() if hasattr(value, 'isoformat') else value

def read_sheet(ws):
    headers = [cell.value for cell in ws[1]]
    rows = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        if not any(value is not None for value in values):
            continue
        rows.append({header: clean(values[i]) for i, header in enumerate(headers) if header})
    return rows


def attach_source_records(rows, ws_sources):
    """把「來源明細」的每來源獨立紀錄（網站＋連結）掛回架上/已下架列。

    讓前端能把同一物件多來源的網頁連結分開顯示（問題五），而不是把
    「台灣房屋/信義房屋」併成單一連結。
    """
    headers = [cell.value for cell in ws_sources[1]]
    sources_by_fingerprint = {}
    for values in ws_sources.iter_rows(min_row=2, values_only=True):
        if not any(value is not None for value in values):
            continue
        d = {header: clean(values[i]) for i, header in enumerate(headers) if header}
        fp = str(d.get('物件指紋', '') or '')
        if not fp:
            continue
        # 已下架的來源記錄代表該仲介的這筆刊登已消失，不應作為「可點」連結。
        if d.get('狀態') == '已下架':
            continue
        sources_by_fingerprint.setdefault(fp, []).append({
            '網站': d.get('來源網站', ''),
            '物件編號': d.get('來源物件編號', ''),
            '連結': d.get('來源連結', ''),
        })
    for row in rows:
        fp = str(row.get('指紋', '') or '')
        if fp not in sources_by_fingerprint:
            continue
        sources = sources_by_fingerprint[fp]
        row['來源物件'] = sources
        # 同步推導顯示欄位，避免徽章仍顯示已下架的來源名卻只有單一連結。
        names = []
        for source in sources:
            name = str(source.get('網站', '') or '')
            if name and name not in names:
                names.append(name)
        if names:
            row['來源網站'] = '/'.join(names)
        preferred = next((s for s in sources if s.get('網站') in ('信義房屋', '永慶房屋') and s.get('連結')), None) \
            or next((s for s in sources if s.get('連結')), None)
        if preferred:
            row['來源連結'] = preferred['連結']
    return rows

def read_source_health():
    try:
        data = json.loads(STATUS_SOURCE.read_text(encoding='utf-8'))
        if isinstance(data, dict) and isinstance(data.get('sources'), dict):
            return data
    except (json.JSONDecodeError, OSError):
        pass
    return None

def main():
    try:
        if not SOURCE.exists():
            raise SystemExit(f'Excel not found: {SOURCE}')
        workbook = load_workbook(SOURCE, data_only=True)
        active_rows = read_sheet(workbook['架上'])
        removed_rows = read_sheet(workbook['已下架'])
        if '來源明細' in workbook.sheetnames:
            active_rows = attach_source_records(active_rows, workbook['來源明細'])
            removed_rows = attach_source_records(removed_rows, workbook['來源明細'])
        payload = {
            'generated_at': datetime.now().isoformat(),
            'source': '本機迴龍物件追蹤.xlsx',
            'active': active_rows,
            'removed': removed_rows,
            'price_changes': read_sheet(workbook['價格變動']),
            'source_health': read_source_health(),
        }
        workbook.close()

        # Avoid a daily Git commit and Vercel deployment when the workbook data is
        # identical. Preserve the previous timestamp so the JSON remains unchanged.
        if DEST.exists():
            try:
                previous = json.loads(DEST.read_text(encoding='utf-8'))
                # Compare only the meaningful data keys
                keys_to_compare = ('source', 'active', 'removed', 'price_changes', 'source_health')
                if all(previous.get(k) == payload.get(k) for k in keys_to_compare):
                    payload['generated_at'] = previous.get('generated_at', payload['generated_at'])
            except (json.JSONDecodeError, OSError):
                pass

        DEST.parent.mkdir(parents=True, exist_ok=True)
        serialized = json.dumps(payload, ensure_ascii=False, indent=2) + '\n'
        if DEST.exists() and DEST.read_text(encoding='utf-8') == serialized:
            print('dashboard data unchanged')
            return
        DEST.write_text(serialized, encoding='utf-8')
        print(f"exported active={len(payload['active'])} removed={len(payload['removed'])} price_changes={len(payload['price_changes'])}")
    except Exception as e:
        print(f"Error during export: {e}", file=sys.stderr)
        sys.exit(1)

if __name__ == '__main__':
    main()